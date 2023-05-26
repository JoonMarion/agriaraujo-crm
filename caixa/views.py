from django.http import HttpResponse
from django.shortcuts import redirect, get_object_or_404, render
from datetime import date, datetime
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from .models import Caixa
import openpyxl
from openpyxl.styles import Font


class TestMixinIsAdmin(UserPassesTestMixin):
    def test_func(self):
        is_admin_or_is_staff = self.request.user.is_superuser or \
                               self.request.user.is_staff
        return bool(is_admin_or_is_staff)

    def handle_no_permission(self):
        messages.error(
            self.request, "Você não tem permissões!"
        )
        return redirect("accounts:index")


class CaixaListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/caixa_list.html'

    def get_queryset(self):
        data_inicio = self.request.GET.get('inicio')
        data_fim = self.request.GET.get('fim')
        if data_inicio and data_fim:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            return Caixa.objects.filter(data__range=(data_inicio, data_fim)).order_by('-data')
        else:
            hoje = date.today()
            return Caixa.objects.filter(data=hoje)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        object_list = context['object_list']
        total_caixa = sum(obj_caixa.quantidade_kg for obj_caixa in object_list)
        data_inicio = self.request.GET.get('inicio', None)
        data_fim = self.request.GET.get('fim', None)
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
        context['data_inicio'] = data_inicio
        context['data_fim'] = data_fim
        context['total_caixa'] = total_caixa
        return context


class CaixaCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Caixa
    login_url = 'accounts:login'
    template_name = 'form_caixa.html'
    fields = ['data', 'descricao', 'quantidade_kg', 'tipo', 'valor_total', 'valor_kg']
    success_url = reverse_lazy('caixa:caixa_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class CaixaUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Caixa
    login_url = 'accounts:login'
    template_name = 'form_caixa.html'
    fields = ['data', 'tipo', 'descricao', 'valor']
    success_url = reverse_lazy('caixa:caixa_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class CaixaDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Caixa
    success_url = reverse_lazy('caixa:caixa_lista')
    template_name = 'form_delete_caixa.html'

    def get_success_url(self):
        messages.success(self.request, "Movimentação excluída com sucesso!")
        return reverse_lazy('caixa:caixa_lista')


def download_xlsx(request, **kwargs):

    data_inicio = request.GET.get('inicio')
    data_fim = request.GET.get('fim')

    if data_inicio and data_fim:
        data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y').date()
        data_fim = datetime.strptime(data_fim, '%d/%m/%Y').date()
        object_list = Caixa.objects.filter(data__range=(data_inicio, data_fim)).order_by('-data')
    else:
        hoje = date.today()
        object_list = Caixa.objects.filter(data=hoje)

    # criar um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caixa"

    # adicionar cabeçalho
    ws['A1'] = 'DATA'
    ws['B1'] = 'TIPO'
    ws['C1'] = 'DESCRIÇÃO'
    ws['D1'] = 'VALOR'

    # adicionar dados
    row_num = 2
    for caixa in object_list:
        ws.cell(row=row_num, column=1, value=caixa.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value=caixa.get_tipo_display())
        ws.cell(row=row_num, column=3, value=caixa.descricao)
        ws.cell(row=row_num, column=4, value=round(caixa.valor, 2))
        row_num += 1

    # Calcular o total dos valores
    total = sum(caixa.valor for caixa in object_list)

    # Adicionar uma linha adicional com o valor total
    ws.cell(row=row_num, column=4, value="Total")
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    row_num += 1
    ws.cell(row=row_num, column=4, value=round(total, 2))

    # configurar o cabeçalho de resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Caixa - {data_inicio} a {data_fim}.xlsx"'

    # salvar o arquivo Excel na resposta
    wb.save(response)

    return response


def caixa_imprimir(request, caixa_id):
    caixa = Caixa.objects.get(id=caixa_id)

    recibo_data = {
        'data': caixa.data.strftime('%d/%m/%Y'),
        'descricao': caixa.descricao,
        'tipo': caixa.tipo,
        'quantidade_kg': caixa.quantidade_kg,
        'valor_kg': caixa.valor_kg,
        'valor_total': caixa.valor_total,
        'recibo_id': caixa_id,
    }

    context = {
        'recibo': recibo_data,
    }

    return render(request, 'receipts/recibo_caixa.html', context)


caixa_cadastro = CaixaCreateView.as_view()
caixa_atualizar = CaixaUpdateView.as_view()
caixa_lista = CaixaListView.as_view()
caixa_deletar = CaixaDeleteView.as_view()
