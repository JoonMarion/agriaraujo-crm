import io
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.generic import CreateView, ListView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from openpyxl.workbook import Workbook

from clientes.models import Cliente, Transacao


class TestMixinIsAdmin(UserPassesTestMixin):
    def test_func(self):
        is_admin_or_is_staff = self.request.user.is_superuser or \
                               self.request.user.is_staff
        return bool(is_admin_or_is_staff)

    def handle_no_permission(self):
        messages.error(
            self.request, "Você não tem permissões!"
        )
        return redirect("accounts:index")


class IndexView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'index.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            queryset = Cliente.objects.filter(nome__icontains=search).order_by('nome')
        else:
            queryset = Cliente.objects.all().order_by('nome')

        # Calculate the sum of quantidade_kg for all clients
        total_quantidade_kg = Transacao.objects.aggregate(Sum('quantidade_kg'))['quantidade_kg__sum']
        # Add the total_quantidade_kg to the context
        self.extra_context = {'total_quantidade_kg': total_quantidade_kg}

        return queryset


class ClienteCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Cliente
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone']
    success_url = reverse_lazy('clientes:cliente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Cliente'
        return context


class ClienteUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Cliente
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone']
    success_url = reverse_lazy('clientes:cliente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Cliente'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class ClienteListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/cliente_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            queryset = Cliente.objects.filter(nome__icontains=search).order_by('nome')
        else:
            queryset = Cliente.objects.all().order_by('nome')

        # Calculate the sum of quantidade_kg for all clients
        total_quantidade_kg = Transacao.objects.aggregate(Sum('quantidade_kg'))['quantidade_kg__sum']
        # Add the total_quantidade_kg to the context
        self.extra_context = {'total_quantidade_kg': total_quantidade_kg}

        return queryset


class ClienteDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Cliente
    success_url = reverse_lazy('clientes:cliente_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Cliente excluído com sucesso!")
        return reverse_lazy('clientes:cliente_lista')


class TransacaoCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Transacao
    login_url = 'accounts:login'
    template_name = 'transacao_form.html'
    fields = ['cliente', 'data', 'descricao', 'quantidade_kg', 'tipo', 'valor_total', 'valor_kg']

    def __init__(self):
        self.success_url = None

    def dispatch(self, request, *args, **kwargs):
        self.success_url = reverse_lazy('clientes:transacao_lista', kwargs={'pk': kwargs['pk']})
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ultimo_parametro = self.request.path.split("/")[-2]
        context['ultimo_parametro'] = ultimo_parametro
        context['form_name'] = 'Transacao'
        form = context.get('form')
        if form:
            form.fields['cliente'].initial = ultimo_parametro

        return context


class TransacaoUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Transacao
    login_url = 'accounts:login'
    template_name = 'transacao_form.html'
    fields = ['cliente', 'data', 'descricao', 'quantidade_kg', 'tipo', 'valor_total', 'valor_kg']

    def __init__(self):
        self.success_url = None

    def dispatch(self, request, *args, **kwargs):
        self.success_url = reverse_lazy('clientes:transacao_lista', kwargs={'pk': kwargs['pk']})
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Transacao'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class TransacaoListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/transacao_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transacao_id = self.kwargs.get('pk')
        total_soma = Transacao.objects.filter(cliente__id=transacao_id).aggregate(total=Sum('quantidade_kg'))['total']
        context['total_soma'] = total_soma
        return context

    def get_queryset(self):
        search = self.request.GET.get('search')
        transacao_id = self.kwargs.get('pk')

        if search:
            return Transacao.objects.filter(descricao__icontains=search, cliente__id=transacao_id).order_by('-data')
        else:
            return Transacao.objects.filter(cliente__id=transacao_id).order_by('-data')


class TransacaoDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Transacao
    success_url = reverse_lazy('clientes:transacao_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Movimentação excluída com sucesso!")
        return reverse_lazy('clientes:transacao_lista', kwargs={'pk': self.kwargs['pk']})


def transacao_imprimir(request, transacao_id):
    transacao = Transacao.objects.get(id=transacao_id)

    recibo_data = {
        'cliente': transacao.cliente,
        'data': transacao.data.strftime('%d/%m/%Y'),
        'descricao': transacao.descricao,
        'tipo': transacao.tipo,
        'quantidade_kg': transacao.quantidade_kg,
        'valor_kg': transacao.valor_kg,
        'valor_total': transacao.valor_total,
        'recibo_id': transacao_id,
    }

    context = {
        'recibo': recibo_data,
    }

    return render(request, 'receipts/recibo_caixa.html', context)


def download_xlsx(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    # Criar um workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"

    # Adicionar cabeçalhos
    headers = ['CLIENTE', 'DATA', 'HISTÓRICO', 'TIPO', 'QUANTIDADE KG', 'PREÇO UNITÁRIO', 'TOTAL',]
    ws.append(headers)

    ws.cell(row=1, column=1, value='CLIENTE')
    ws.cell(row=2, column=1, value=cliente.nome)

    # Adicionar dados
    saldo = 0
    for transacao in Transacao.objects.select_related('cliente').filter(cliente__id=cliente_id):
        cliente = transacao.cliente.nome
        data = transacao.data.strftime('%d/%m/%Y')
        descricao = transacao.descricao
        tipo = 'Entrada' if transacao.tipo == 'E' else 'Saída'
        quantidade_kg = str(transacao.quantidade_kg or "")
        valor_kg = "" if transacao.tipo == 'E' else 'R$ ' + str(transacao.valor_kg)
        valor_total = "" if transacao.tipo == 'E' else 'R$ ' + str(transacao.valor_total or "")
        row = [cliente, data, descricao, tipo, quantidade_kg, valor_kg, valor_total]
        ws.append(row)
        saldo += transacao.quantidade_kg

    # Adicionar campo "SALDO"
    ws.append(['', '', '', '', '', '', ''])
    ws.append(['', '', '', '', '', 'SALDO', saldo])

    # Configurar a resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Movimentações de {cliente}.xlsx"'

    # Salvar o arquivo XLSX no buffer de memória
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Copiar o conteúdo do buffer para a resposta
    response.write(buffer.getvalue())

    return response



transacao_cadastro = TransacaoCreateView.as_view()
transacao_atualizar = TransacaoUpdateView.as_view()
transacao_lista = TransacaoListView.as_view()
transacao_deletar = TransacaoDeleteView.as_view()

cliente_cadastro = ClienteCreateView.as_view()
cliente_atualizar = ClienteUpdateView.as_view()
cliente_lista = ClienteListView.as_view()
cliente_deletar = ClienteDeleteView.as_view()
