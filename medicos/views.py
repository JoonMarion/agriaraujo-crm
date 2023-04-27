from datetime import date, datetime
from django.db.models import Q
from django.http import request
from django.views.generic import CreateView, ListView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import Medico, Agenda, Especialidade, Cliente, Relatorio, Caixa
import openpyxl
from django.http import HttpResponse


class TestMixinIsAdmin(UserPassesTestMixin):
    def test_func(self):
        is_admin_or_is_staff = self.request.user.is_superuser or \
                               self.request.user.is_staff
        return bool(is_admin_or_is_staff)

    def handle_no_permission(self):
        messages.error(
            self.request, "Você não tem permissões!"
        )
        return redirect("accounts:index")


class IndexView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'index.html'

    today = date.today()

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Agenda.objects.filter(dia=self.today).filter(Q(dia__icontains=search) | Q(cliente__nome__icontains=search) | Q(
                medico__nome__icontains=search)).order_by('dia')
        else:
            return Agenda.objects.filter(data=self.today).order_by('horario')


class MedicoCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Medico
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone']
    success_url = reverse_lazy('medicos:medicos_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Profissional'
        return context


class MedicoUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Medico
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone']
    success_url = reverse_lazy('medicos:medicos_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Profissional'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class MedicoListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/medicos_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Medico.objects.filter(nome__icontains=search).order_by('nome')
        else:
            return Medico.objects.all().order_by('nome')


class MedicoDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Medico
    success_url = reverse_lazy('medicos:medicos_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Profissional excluído com sucesso!")
        return reverse_lazy('medicos:medicos_lista')


class EspecialidadeCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Especialidade
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome']
    success_url = reverse_lazy('medicos:especialidade_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Procedimento'
        return context


class EspecialidadeUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Especialidade
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome']
    success_url = reverse_lazy('medicos:especialidade_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Procedimento'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class EspecialidadeListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/especialidade_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Especialidade.objects.filter(nome__icontains=search).order_by('nome')
        else:
            return Especialidade.objects.all().order_by('nome')


class EspecialidadeDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Especialidade
    success_url = reverse_lazy('medicos:especialidade_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Procedimento excluído com sucesso!")
        return reverse_lazy('medicos:especialidade_lista')


class ClienteCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Cliente
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone']
    success_url = reverse_lazy('medicos:cliente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Cliente'
        return context


class ClienteUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Cliente
    login_url = 'accounts:login'
    template_name = 'form.html'
    fields = ['nome', 'telefone', 'avaliacao']
    success_url = reverse_lazy('medicos:cliente_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Cliente'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class ClienteDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Cliente
    success_url = reverse_lazy('medicos:cliente_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Cliente excluído com sucesso!")
        return reverse_lazy('medicos:cliente_lista')


class ClienteListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/cliente_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Cliente.objects.filter(nome__icontains=search).order_by('nome')
        else:
            return Cliente.objects.all().order_by('nome')




class AgendaCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Agenda
    login_url = 'accounts:login'
    template_name = 'form_agenda.html'
    fields = ['medico', 'data', 'horario', 'cliente', 'procedimento']
    success_url = reverse_lazy('medicos:agenda_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_name'] = 'Agenda'
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class AgendaUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Agenda
    login_url = 'accounts:login'
    template_name = 'form_agenda.html'
    fields = ['medico', 'data', 'horario', 'cliente', 'procedimento']
    success_url = reverse_lazy('medicos:agenda_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class AgendaDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Agenda
    success_url = reverse_lazy('medicos:agenda_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Consulta excluída com sucesso!")
        return reverse_lazy('medicos:agenda_lista')


class AgendaListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/agenda_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Agenda.objects.filter(Q(dia__icontains=search) | Q(cliente__nome__icontains=search) | Q(medico__nome__icontains=search)).order_by('data')
        else:
            return Agenda.objects.all().order_by('data')


class RelatorioCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Relatorio
    login_url = 'accounts:login'
    template_name = 'form_relatorio.html'
    fields = ['data', 'relatorio']
    success_url = reverse_lazy('medicos:relatorio_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class RelatorioUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Relatorio
    login_url = 'accounts:login'
    template_name = 'form_relatorio.html'
    fields = ['data', 'relatorio']
    success_url = reverse_lazy('medicos:relatorio_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class RelatorioListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/relatorio_list.html'

    def get_queryset(self):
        search = self.request.GET.get('search')
        if search:
            return Relatorio.objects.filter(data__icontains=search).order_by('-data')
        else:
            return Relatorio.objects.all().order_by('-data')


class RelatorioDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Relatorio
    success_url = reverse_lazy('medicos:relatorio_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Relatório excluído com sucesso!")
        return reverse_lazy('medicos:relatorio_lista')


class CaixaListView(LoginRequiredMixin, TestMixinIsAdmin, ListView):
    login_url = 'accounts:login'
    template_name = 'lists/caixa_list.html'

    def get_queryset(self):
        data_inicio = self.request.GET.get('inicio')
        data_fim = self.request.GET.get('fim')
        if data_inicio and data_fim:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            return Caixa.objects.filter(data__range=(data_inicio, data_fim)).order_by('-data')
        else:
            hoje = date.today()
            return Caixa.objects.filter(data=hoje)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        object_list = context['object_list']
        total = sum([obj.valor for obj in object_list])
        data_inicio = self.request.GET.get('inicio', None)
        data_fim = self.request.GET.get('fim', None)
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
        context['data_inicio'] = data_inicio
        context['data_fim'] = data_fim
        context['total'] = total
        return context


def download_xlsx(request, **kwargs):

    data_inicio = request.GET.get('inicio')
    data_fim = request.GET.get('fim')

    if data_inicio and data_fim:
        data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y').date()
        data_fim = datetime.strptime(data_fim, '%d/%m/%Y').date()
        object_list = Caixa.objects.filter(data__range=(data_inicio, data_fim)).order_by('-data')
    else:
        hoje = date.today()
        object_list = Caixa.objects.filter(data=hoje)

    # criar um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caixa"

    # adicionar cabeçalho
    ws['A1'] = 'Data'
    ws['B1'] = 'Tipo'
    ws['C1'] = 'Descrição'
    ws['D1'] = 'Valor'

    # adicionar dados
    row_num = 2
    for caixa in object_list:
        ws.cell(row=row_num, column=1, value=caixa.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value=caixa.get_tipo_display())
        ws.cell(row=row_num, column=3, value=caixa.descricao)
        if caixa.tipo == 'E':
            ws.cell(row=row_num, column=4, value=caixa.valor).number_format = '#,##0.00 [$R$]'
        else:
            ws.cell(row=row_num, column=4, value=-caixa.valor).number_format = '-#,##0.00 [$R$]'
        row_num += 1

    # configurar o cabeçalho de resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Caixa - {data_inicio} a {data_fim}.xlsx"'

    # salvar o arquivo Excel na resposta
    wb.save(response)

    return response


class CaixaCreateView(LoginRequiredMixin, TestMixinIsAdmin, CreateView):
    model = Caixa
    login_url = 'accounts:login'
    template_name = 'form_caixa.html'
    fields = ['data', 'tipo', 'descricao', 'valor']
    success_url = reverse_lazy('medicos:caixa_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class CaixaUpdateView(LoginRequiredMixin, TestMixinIsAdmin, UpdateView):
    model = Caixa
    login_url = 'accounts:login'
    template_name = 'form_caixa.html'
    fields = ['data', 'tipo', 'descricao', 'valor']
    success_url = reverse_lazy('medicos:caixa_lista')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class CaixaDeleteView(LoginRequiredMixin, TestMixinIsAdmin, DeleteView):
    model = Caixa
    success_url = reverse_lazy('medicos:caixa_lista')
    template_name = 'form_delete.html'

    def get_success_url(self):
        messages.success(self.request, "Transação excluída com sucesso!")
        return reverse_lazy('medicos:caixa_lista')


medico_cadastro = MedicoCreateView.as_view()
medico_atualizar = MedicoUpdateView.as_view()
medico_lista = MedicoListView.as_view()
medico_deletar = MedicoDeleteView.as_view()

especialidade_cadastro = EspecialidadeCreateView.as_view()
especialidade_atualizar = EspecialidadeUpdateView.as_view()
especialidade_lista = EspecialidadeListView.as_view()
especialidade_deletar = EspecialidadeDeleteView.as_view()

cliente_cadastro = ClienteCreateView.as_view()
cliente_atualizar = ClienteUpdateView.as_view()
cliente_lista = ClienteListView.as_view()
cliente_deletar = ClienteDeleteView.as_view()

agenda_cadastro = AgendaCreateView.as_view()
agenda_atualizar = AgendaUpdateView.as_view()
agenda_lista = AgendaListView.as_view()
agenda_deletar = AgendaDeleteView.as_view()

relatorio_cadastro = RelatorioCreateView.as_view()
relatorio_atualizar = RelatorioUpdateView.as_view()
relatorio_lista = RelatorioListView.as_view()
relatorio_deletar = RelatorioDeleteView.as_view()

caixa_cadastro = CaixaCreateView.as_view()
caixa_atualizar = CaixaUpdateView.as_view()
caixa_lista = CaixaListView.as_view()
caixa_deletar = CaixaDeleteView.as_view()